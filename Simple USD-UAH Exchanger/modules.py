import openpyxl
import csv
import pprint


def data_csv_read(filename: str) -> dict:
    """
    This function reads data from csv-file and returns it as a dictionary.
    :param filename: str - name of file with system data
    :return: dict - system data
    """
    with open(filename, 'r', encoding='utf-8') as file:
        system_data = {}
        reader = csv.DictReader(file)
        for row in reader:
            system_data[row['CURRENCY']] = {
                'RATE': float(row['RATE']),
                'AVAILABLE': float(row['AVAILABLE'])
            }
    return system_data


def data_csv_write(system_data: dict, filename: str):
    """
    This function writes data from dictionary to csv-file.
    :param system_data: dict - system data
    :param filename: str - name of file you want to save data into
    :return: None
    """
    with open(filename, 'w', newline='', encoding='utf-8') as file:
        fieldnames = ['CURRENCY', 'RATE', 'AVAILABLE']
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        for key in system_data.keys():
            writer.writerow(
                {'CURRENCY': key, 'RATE': system_data[key]['RATE'], 'AVAILABLE': system_data[key]['AVAILABLE']})


def data_xlsx_read(filename: str) -> dict:
    """
    This function reads data from xlsx-file and returns it as a dictionary.
    :param filename: str - name of file with system data
    :return: dict - system data
    """
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active

    system_data = {}
    for i in range(2, sheet.max_row + 1):
        system_data[sheet.cell(row=i, column=1).value] = {
            'RATE': float(sheet.cell(row=i, column=2).value),
            'AVAILABLE': float(sheet.cell(row=i, column=3).value)
        }
    return system_data


def data_xlsx_write(system_data, filename):
    """
    This function saves changed system data to xlsx-file.
    :param system_data: dict - system data with changed values that needs to be saved
    :param filename: str - name of file you want to save data into
    :return: None
    """
    book = openpyxl.Workbook()
    sheet = book.active
    sheet['A1'] = 'CURRENCY'
    sheet['B1'] = 'RATE'
    sheet['C1'] = 'AVAILABLE'
    sheet['A2'] = 'USD'
    sheet['B2'] = system_data['USD']['RATE']
    sheet['C2'] = system_data['USD']['AVAILABLE']
    sheet['A3'] = 'UAH'
    sheet['B3'] = system_data['UAH']['RATE']
    sheet['C3'] = system_data['UAH']['AVAILABLE']
    book.save(filename)
    book.close()


def possible_actions(actions_list: list):
    """
    This function prints actions that user can manipulate.
    :param actions_list: list - list of possible actions
    :return: None
    """
    for i in range(len(actions_list)):
        print(f"> {i + 1}. {actions_list[i]}")


def course(system_data: dict, currency: list[str]) -> str:
    """
    This function returns course of certain input currency.
    :param system_data: dict - system data
    :param currency: list[str] - currency you want to know the course of
    :return: str - course of currency
    """
    if currency[1] in system_data.keys():
        return f"RATE {system_data[currency[1]]['RATE']}, AVAILABLE {system_data[currency[1]]['AVAILABLE']}"
    else:
        return f"INVALID CURRENCY {currency[1]}"


def system_data_change(sum: float, result: float, transfer_from: str, transfer_to: str, system_data: dict) -> dict:
    """
    This function changes system data after exchange.
    :param sum: float - amount of exchange
    :param result: float - amount of exchange after rounding
    :param transfer_from: str - currency you want to exchange from
    :param transfer_to: str - currency you want to exchange to
    :param system_data: dict - system data
    :return: dict - changed dictionary with a system data
    """
    system_data[transfer_from]['AVAILABLE'] = round(system_data[transfer_from]['AVAILABLE'] - result, 2)
    system_data[transfer_to]['AVAILABLE'] = round(system_data[transfer_to]['AVAILABLE'] + sum, 2)

    return system_data


def exchange(summ: float, transfer_from: str, transfer_to: str, system_data: dict) -> str:
    """
    This function exchanges currencies.
    :param summ: float - amount of exchange
    :param transfer_from: str - currency you want to exchange from
    :param transfer_to: str - currency you want to exchange to
    :param system_data: dict - system data
    :return: str - result of exchange
    """
    herf = summ / system_data[transfer_from]['RATE']
    result = round(herf, 2)
    if result > system_data[transfer_from]['AVAILABLE']:
        return f"UNAVAILABLE, REQUIRED BALANCE {result}, AVAILABLE {system_data[transfer_from]['AVAILABLE']}"
    else:
        # pprint.pprint(system_data_change(summ, result, transfer_from, transfer_to, system_data))
        system_data_change(summ, result, transfer_from, transfer_to, system_data)
        return f"{transfer_from} {result}, RATE {system_data[transfer_to]['RATE']}"
